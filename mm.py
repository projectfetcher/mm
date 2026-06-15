#!/usr/bin/env python3
"""
MIMU Jobs Scraper
=================
Scrapes ALL job listings from https://themimu.info/jobs-for-myanmar-nationals
Downloads each PDF, extracts text, then uses regex/keyword parsing to extract
structured job fields — saving everything to an Excel file.

NO API KEY REQUIRED. No external AI services used.

Requirements:
    pip install requests pdfplumber pandas openpyxl beautifulsoup4

Usage:
    python mimu_jobs_scraper.py

Output:
    mimu_jobs_YYYYMMDD.xlsx  — structured job data
    pdfs/                    — downloaded PDF files
"""

import os
import re
import time
import datetime
import requests
import pdfplumber
import pandas as pd
from pathlib import Path
from bs4 import BeautifulSoup

# ── Configuration ──────────────────────────────────────────────────────────────
BASE_URL   = "https://themimu.info"
JOBS_URL   = f"{BASE_URL}/jobs-for-myanmar-nationals"
PDF_DIR    = Path("pdfs")
OUTPUT_XLS = f"mimu_jobs_{datetime.date.today():%Y%m%d}.xlsx"

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.5",
    "Accept-Encoding": "gzip, deflate, br",
    "Connection": "keep-alive",
    "Upgrade-Insecure-Requests": "1",
}

OUTPUT_COLUMNS = [
    "Job Title", "Job Type", "Job Qualifications", "Job Experience",
    "Job Location", "Job Field", "Date Posted", "Deadline",
    "Job Description", "Application", "Company URL", "Company Name",
    "Company Logo", "Company Industry", "Company Founded", "Company Type",
    "Company Website", "Company Address", "Company Details",
    "Job URL", "Estimated Deadline", "Salary Range",
]

# ── Sector / field keyword map ─────────────────────────────────────────────────
FIELD_KEYWORDS = {
    "Health":         ["health", "medical", "doctor", "nurse", "clinical", "pharmacy", "nutrition", "epidemic", "disease"],
    "Logistics":      ["logistics", "supply chain", "fleet", "transport", "warehouse", "procurement"],
    "Finance":        ["finance", "accounting", "audit", "budget", "financial", "grants", "treasury"],
    "WASH":           ["wash", "water", "sanitation", "hygiene", "latrine"],
    "Protection":     ["protection", "gbv", "gender", "child protection", "safeguarding", "legal aid"],
    "Education":      ["education", "teacher", "school", "training", "learning", "teacher"],
    "HR":             ["human resource", "hr ", "recruitment", "personnel", "talent"],
    "Administration": ["admin", "administration", "office management", "receptionist"],
    "Food Security":  ["food security", "livelihoods", "agriculture", "food assistance"],
    "Shelter":        ["shelter", "nfi", "construction", "engineer", "infrastructure"],
    "CCCM":           ["cccm", "camp", "site management"],
    "Nutrition":      ["nutrition", "malnutrition", "stunting", "wasting", "feeding"],
    "IT":             ["it ", "information technology", "software", "developer", "database", "network", "ict"],
    "Communications": ["communication", "media", "journalist", "reporting", "public relations", "advocacy"],
    "Legal":          ["legal", "lawyer", "compliance", "policy"],
}

JOB_TYPE_KEYWORDS = {
    "Consultancy":  ["consultanc", "consultant", "consultancy", "TOR", "terms of reference"],
    "Internship":   ["intern", "internship"],
    "Part-time":    ["part-time", "part time"],
    "Contract":     ["contract", "fixed-term", "fixed term", "temporary"],
    "Full-time":    ["full-time", "full time", "permanent", "national staff"],
}


# ── Helpers ────────────────────────────────────────────────────────────────────

def fetch_page(url: str, retries: int = 5) -> str:
    for attempt in range(1, retries + 1):
        try:
            r = requests.get(url, headers=HEADERS, timeout=60)
            r.raise_for_status()
            return r.text
        except Exception as e:
            wait = 3 * attempt
            print(f"  [!] Attempt {attempt}/{retries} failed: {e}")
            if attempt < retries:
                print(f"      Retrying in {wait}s …")
                time.sleep(wait)
    return ""


def download_pdf(url: str, dest: Path) -> bool:
    if dest.exists() and dest.stat().st_size > 0:
        return True
    try:
        r = requests.get(url, headers=HEADERS, timeout=120, stream=True)
        r.raise_for_status()
        dest.write_bytes(r.content)
        return True
    except Exception as e:
        print(f"  [!] PDF download failed ({url}): {e}")
        return False


def extract_pdf_text(pdf_path: Path) -> str:
    try:
        with pdfplumber.open(pdf_path) as pdf:
            parts = []
            for page in pdf.pages:
                t = page.extract_text()
                if t:
                    parts.append(t.strip())
            return "\n".join(parts)
    except Exception as e:
        print(f"  [!] PDF text extraction failed ({pdf_path}): {e}")
        return ""


# ── Table parser ───────────────────────────────────────────────────────────────

def parse_jobs_table(html: str) -> list:
    soup = BeautifulSoup(html, "html.parser")
    jobs = []

    job_table = None
    for table in soup.find_all("table"):
        header_row = table.find("tr")
        if header_row:
            header_text = header_row.get_text(" ", strip=True).lower()
            if "closing date" in header_text or "job title" in header_text:
                job_table = table
                break

    if not job_table:
        for table in soup.find_all("table"):
            if len(table.find_all("tr")) >= 5:
                job_table = table
                break

    if not job_table:
        print("[!] Could not locate the jobs table in the HTML.")
        return jobs

    rows = job_table.find_all("tr")
    print(f"      Table rows found (inc. header): {len(rows)}")

    col_map = {}
    header_row = rows[0]
    for idx, cell in enumerate(header_row.find_all(["th", "td"])):
        text = cell.get_text(" ", strip=True).lower()
        if "job title" in text:
            col_map["title"] = idx
        elif "description" in text or "download" in text:
            col_map["pdf"] = idx
        elif "application" in text or "online" in text:
            col_map["app"] = idx
        elif "location" in text or "post" in text:
            col_map["location"] = idx
        elif "organisation" in text or "organization" in text:
            col_map["org"] = idx
        elif "upload" in text or "posted" in text or "date of" in text:
            col_map["date_posted"] = idx
        elif "closing" in text or "deadline" in text:
            col_map["deadline"] = idx
        elif "remark" in text:
            col_map["remarks"] = idx

    defaults = {
        "title": 0, "pdf": 1, "app": 2,
        "location": 3, "org": 4, "date_posted": 5,
        "deadline": 6, "remarks": 7,
    }
    for k, v in defaults.items():
        col_map.setdefault(k, v)

    for row in rows[1:]:
        cols = row.find_all(["td", "th"])
        if not cols:
            continue

        def get_col(key):
            idx = col_map.get(key, -1)
            return cols[idx] if 0 <= idx < len(cols) else None

        title_cell    = get_col("title")
        pdf_cell      = get_col("pdf")
        app_cell      = get_col("app")
        loc_cell      = get_col("location")
        org_cell      = get_col("org")
        posted_cell   = get_col("date_posted")
        deadline_cell = get_col("deadline")
        remarks_cell  = get_col("remarks")

        if not title_cell:
            continue
        title = title_cell.get_text(strip=True)
        if not title or title.lower() in ("job title", "title"):
            continue

        pdf_url = ""
        if pdf_cell:
            dl_link = pdf_cell.find("a", href=True)
            if dl_link:
                href = dl_link["href"]
                if href.startswith("http"):
                    pdf_url = href
                elif href.startswith("//"):
                    pdf_url = "https:" + href
                else:
                    pdf_url = BASE_URL.rstrip("/") + "/" + href.lstrip("/")

        app_url = ""
        if app_cell:
            app_link = app_cell.find("a", href=True)
            if app_link:
                app_url = app_link["href"]
            else:
                raw = app_cell.get_text(strip=True)
                if raw.startswith("http") or "@" in raw:
                    app_url = raw

        location    = loc_cell.get_text("  ", strip=True)  if loc_cell      else ""
        org         = org_cell.get_text(strip=True)         if org_cell      else ""
        date_posted = posted_cell.get_text(strip=True)      if posted_cell   else ""
        deadline    = deadline_cell.get_text(strip=True)    if deadline_cell else ""
        remarks     = remarks_cell.get_text(strip=True)     if remarks_cell  else ""

        location = re.sub(r"\s{2,}", ", ", location).strip(", ")

        jobs.append({
            "title":       title,
            "pdf_url":     pdf_url,
            "app_url":     app_url,
            "location":    location,
            "org":         org,
            "date_posted": date_posted,
            "deadline":    deadline,
            "remarks":     remarks,
            "job_url":     JOBS_URL,
        })

    return jobs


# ── Regex-based field extractor (replaces Claude AI) ──────────────────────────

def _search(patterns: list, text: str, flags=re.IGNORECASE) -> str:
    """Return first non-empty match group from a list of regex patterns."""
    for pattern in patterns:
        m = re.search(pattern, text, flags)
        if m:
            result = m.group(1).strip(" :.,-\t")
            if result:
                return result
    return ""


def _search_block(header_patterns: list, text: str, max_lines: int = 6) -> str:
    """
    Find a section header then grab the next max_lines lines as content.
    Good for multi-line fields like Job Description or Qualifications.
    """
    lines = text.splitlines()
    for i, line in enumerate(lines):
        for pat in header_patterns:
            if re.search(pat, line, re.IGNORECASE):
                block_lines = []
                for j in range(i + 1, min(i + 1 + max_lines, len(lines))):
                    l = lines[j].strip()
                    if not l:
                        continue
                    # Stop if we hit another section header (ALL CAPS or ends with colon)
                    if re.match(r'^[A-Z][A-Z\s]{4,}:?\s*$', l) or (l.endswith(":") and len(l) < 50):
                        break
                    block_lines.append(l)
                if block_lines:
                    return " ".join(block_lines)
    return ""


def detect_job_field(text: str) -> str:
    text_lower = text.lower()
    for field, keywords in FIELD_KEYWORDS.items():
        if any(kw in text_lower for kw in keywords):
            return field
    return "Other"


def detect_job_type(text: str) -> str:
    text_lower = text.lower()
    for jtype, keywords in JOB_TYPE_KEYWORDS.items():
        if any(kw in text_lower for kw in keywords):
            return jtype
    return "Full-time"


def extract_salary(text: str) -> str:
    patterns = [
        r'salary[:\s]+([^\n]{5,60})',
        r'remuneration[:\s]+([^\n]{5,60})',
        r'compensation[:\s]+([^\n]{5,60})',
        r'([\$MMK][\d,]+(?:\s*[-–]\s*[\$MMK][\d,]+)?(?:\s*(?:per\s+month|/month|monthly|annually))?)',
        r'(\d[\d,]+\s*(?:MMK|USD|Ks|Kyats)(?:\s*[-–]\s*\d[\d,]+\s*(?:MMK|USD|Ks|Kyats))?)',
    ]
    return _search(patterns, text)


def extract_experience(text: str) -> str:
    patterns = [
        r'experience[:\s]+([^\n]{5,80})',
        r'(\d+\+?\s*(?:to\s*\d+\s*)?years?[^\n]{0,50}experience[^\n]{0,30})',
        r'(minimum\s+\d+\s+years?[^\n]{0,50})',
        r'(at least\s+\d+\s+years?[^\n]{0,50})',
    ]
    return _search(patterns, text)


def extract_qualifications(text: str) -> str:
    # First try a block search
    block = _search_block(
        [r'qualif', r'education', r'academic', r'degree required'],
        text, max_lines=5
    )
    if block:
        return block[:300]
    # Fallback inline
    patterns = [
        r'(?:qualifications?|education)[:\s]+([^\n]{10,200})',
        r"(bachelor'?s?|master'?s?|phd|diploma|degree)[^\n]{0,100}",
    ]
    return _search(patterns, text)[:300]


def extract_description(text: str) -> str:
    """Extract a 2-4 sentence summary from the most descriptive part of the PDF."""
    # Look for a responsibilities / duties section
    block = _search_block(
        [r'responsibilit', r'duties', r'key tasks', r'scope of work', r'objective'],
        text, max_lines=8
    )
    if block:
        # Trim to ~400 chars
        return block[:400]

    # Fallback: take the meatiest paragraph (longest line cluster)
    lines = [l.strip() for l in text.splitlines() if len(l.strip()) > 40]
    if lines:
        return " ".join(lines[:4])[:400]
    return text[:400]


def extract_company_website(text: str, app_url: str) -> str:
    patterns = [
        r'website[:\s]+(https?://[^\s]+)',
        r'(https?://(?!.*apply|.*career|.*job|.*lever|.*greenhouse|.*workable)[^\s]{10,})',
    ]
    result = _search(patterns, text)
    if not result and app_url.startswith("http"):
        # Use domain of app_url as a fallback
        m = re.match(r'(https?://[^/]+)', app_url)
        if m:
            return m.group(1)
    return result


def extract_address(text: str) -> str:
    patterns = [
        r'address[:\s]+([^\n]{10,120})',
        r'office[:\s]+([^\n]{10,120})',
        r'headquarter[s]?[:\s]+([^\n]{10,80})',
    ]
    return _search(patterns, text)


def extract_company_details(text: str, org: str) -> str:
    block = _search_block(
        [r'about\s+(?:us|the\s+organization|' + re.escape(org[:10]) + r')',
         r'background', r'organization overview'],
        text, max_lines=6
    )
    return block[:400] if block else ""


def parse_fields(raw_text: str, meta: dict) -> dict:
    """
    Pure regex/keyword extraction — no API required.
    Falls back to table metadata where PDF text is unavailable.
    """
    text = raw_text or ""
    org  = meta.get("org", "")

    job_title       = meta.get("title", "") or _search([r'position[:\s]+([^\n]{3,80})', r'job title[:\s]+([^\n]{3,80})'], text)
    job_location    = meta.get("location", "") or _search([r'location[:\s]+([^\n]{3,80})', r'duty station[:\s]+([^\n]{3,80})'], text)
    date_posted     = meta.get("date_posted", "")
    deadline        = meta.get("deadline", "") or _search([
        r'closing date[:\s]+([^\n]{3,40})',
        r'deadline[:\s]+([^\n]{3,40})',
        r'application deadline[:\s]+([^\n]{3,40})',
    ], text)

    job_type        = detect_job_type(text + " " + job_title)
    job_field       = detect_job_field(text + " " + job_title)
    qualifications  = extract_qualifications(text)
    experience      = extract_experience(text)
    description     = extract_description(text)
    salary          = extract_salary(text)
    app_url         = meta.get("app_url", "") or _search([r'apply[:\s]+(https?://[^\s]+)', r'application[:\s]+(https?://[^\s]+)'], text)
    company_website = extract_company_website(text, app_url)
    company_address = extract_address(text)
    company_details = extract_company_details(text, org)

    # Company type heuristics
    company_type = ""
    tl = text.lower()
    if any(w in tl for w in ["ngo", "non-governmental", "non governmental"]):
        company_type = "NGO"
    elif any(w in tl for w in ["united nations", " un ", "undp", "unicef", "wfp", "unhcr", "who ", "ilo "]):
        company_type = "UN Agency"
    elif any(w in tl for w in ["government", "ministry", "department of"]):
        company_type = "Government"
    elif any(w in tl for w in ["private", "company", "ltd", "limited", "corporation"]):
        company_type = "Private Sector"

    # Company industry — same as Job Field for NGO context
    company_industry = job_field

    return {
        "Job Title":          job_title,
        "Job Type":           job_type,
        "Job Qualifications": qualifications,
        "Job Experience":     experience,
        "Job Location":       job_location,
        "Job Field":          job_field,
        "Date Posted":        date_posted,
        "Deadline":           deadline,
        "Job Description":    description,
        "Application":        app_url,
        "Company URL":        company_website,
        "Company Name":       org,
        "Company Logo":       "",
        "Company Industry":   company_industry,
        "Company Founded":    "",
        "Company Type":       company_type,
        "Company Website":    company_website,
        "Company Address":    company_address,
        "Company Details":    company_details,
        "Job URL":            meta.get("job_url", ""),
        "Estimated Deadline": deadline,
        "Salary Range":       salary,
    }


# ── Excel writer ───────────────────────────────────────────────────────────────

def save_excel(records: list, path: str):
    df = pd.DataFrame(records, columns=OUTPUT_COLUMNS)

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Jobs")
        ws = writer.sheets["Jobs"]

        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        header_fill = PatternFill(fill_type="solid", fgColor="1F4E79")
        header_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        thin = Border(
            left=Side(style="thin"),  right=Side(style="thin"),
            top=Side(style="thin"),   bottom=Side(style="thin"),
        )

        for col_idx, col_name in enumerate(OUTPUT_COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill      = header_fill
            cell.font      = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border    = thin

        data_font = Font(name="Arial", size=9)
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.font      = data_font
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border    = thin

        col_widths = {
            "Job Title": 35, "Job Type": 15, "Job Qualifications": 40,
            "Job Experience": 25, "Job Location": 25, "Job Field": 18,
            "Date Posted": 14, "Deadline": 14, "Job Description": 60,
            "Application": 35, "Company URL": 30, "Company Name": 25,
            "Company Logo": 15, "Company Industry": 20, "Company Founded": 15,
            "Company Type": 15, "Company Website": 30, "Company Address": 30,
            "Company Details": 35, "Job URL": 40, "Estimated Deadline": 18,
            "Salary Range": 20,
        }
        for col_idx, col_name in enumerate(OUTPUT_COLUMNS, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(col_name, 20)

        ws.row_dimensions[1].height = 30
        ws.freeze_panes = "A2"

    print(f"\n✅  Saved {len(records)} jobs → {path}")


# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    PDF_DIR.mkdir(exist_ok=True)

    print(f"[1/4] Fetching jobs page: {JOBS_URL}")
    html = fetch_page(JOBS_URL)
    if not html:
        print("ERROR: Could not fetch the jobs page. Check your network and try again.")
        return

    print("[2/4] Parsing jobs table …")
    jobs = parse_jobs_table(html)
    print(f"      Found {len(jobs)} job listings.")

    if not jobs:
        print("No jobs found — the page structure may have changed.")
        debug_path = Path("debug_page.html")
        debug_path.write_text(html[:5000], encoding="utf-8")
        print(f"      First 5000 chars saved to {debug_path} for inspection.")
        return

    records = []
    print("[3/4] Downloading PDFs and extracting data …")

    for i, job in enumerate(jobs, start=1):
        title   = job["title"]
        pdf_url = job["pdf_url"]
        print(f"\n  [{i}/{len(jobs)}] {title}")

        raw_text = ""
        if pdf_url:
            url_slug = re.sub(r"[^\w\-.]", "_", pdf_url.split("/")[-1])[:80]
            if not url_slug.lower().endswith(".pdf"):
                url_slug += ".pdf"
            pdf_path = PDF_DIR / url_slug

            ok = download_pdf(pdf_url, pdf_path)
            if ok:
                raw_text = extract_pdf_text(pdf_path)
                print(f"      PDF: {len(raw_text)} chars extracted from {pdf_path.name}")
            else:
                print("      PDF download failed — using metadata only")
        else:
            print("      No PDF link — using metadata only")

        extracted = parse_fields(raw_text, job)
        row = {col: extracted.get(col, "") for col in OUTPUT_COLUMNS}
        records.append(row)

        time.sleep(0.3)

    print("\n[4/4] Saving to Excel …")
    save_excel(records, OUTPUT_XLS)
    print(f"\nDone! Open '{OUTPUT_XLS}' to view all {len(records)} jobs.\n")


if __name__ == "__main__":
    main()
